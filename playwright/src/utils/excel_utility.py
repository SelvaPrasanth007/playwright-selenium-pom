import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from typing import List, Dict, Any, Union
import os


class ExcelUtility:
    """Utility class for reading and writing Excel files"""

    def __init__(self, file_path: str):
        """Initialize with file path"""
        self.file_path = file_path

    def read_excel_file(self, sheet_name: str = 'Sheet1') -> List[Dict[str, Any]]:
        """
        Read data from an Excel file
        
        Args:
            sheet_name: Name of the sheet to read from
            
        Returns:
            List of dictionaries containing the data
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f'Excel file not found at: {self.file_path}')

        workbook = openpyxl.load_workbook(self.file_path)
        
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f'Sheet "{sheet_name}" not found in Excel file')

        worksheet = workbook[sheet_name]
        data = []

        # Get headers from first row
        headers = []
        for cell in worksheet[1]:
            if cell.value:
                headers.append(cell.value)

        # Get data from remaining rows
        for row in worksheet.iter_rows(min_row=2, values_only=False):
            row_data = {}
            for idx, cell in enumerate(row):
                if idx < len(headers):
                    row_data[headers[idx]] = cell.value
            if any(row_data.values()):  # Only add if row has data
                data.append(row_data)

        workbook.close()
        return data

    def read_all_sheets(self) -> Dict[str, List[Dict[str, Any]]]:
        """
        Read all sheets from an Excel file
        
        Returns:
            Dictionary with sheet names as keys and data arrays as values
        """
        if not os.path.exists(self.file_path):
            raise FileNotFoundError(f'Excel file not found at: {self.file_path}')

        workbook = openpyxl.load_workbook(self.file_path)
        all_data = {}

        for sheet_name in workbook.sheetnames:
            all_data[sheet_name] = self.read_excel_file(sheet_name)

        workbook.close()
        return all_data

    def write_excel_file(self, data: List[Dict[str, Any]], sheet_name: str = 'Sheet1') -> None:
        """
        Write data to an Excel file
        
        Args:
            data: List of dictionaries to write
            sheet_name: Name of the sheet to write to
        """
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        if not data:
            workbook.save(self.file_path)
            return

        # Write headers
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_idx, value=header)

        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                worksheet.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        workbook.save(self.file_path)

    def get_row_data(self, sheet_name: str, row_index: int) -> Union[Dict[str, Any], None]:
        """
        Get specific row data from Excel
        
        Args:
            sheet_name: Name of the sheet
            row_index: Index of the row (0-based)
            
        Returns:
            Dictionary with row data or None if row doesn't exist
        """
        data = self.read_excel_file(sheet_name)
        return data[row_index] if row_index < len(data) else None

    def get_column_data(self, sheet_name: str, column_name: str) -> List[Union[str, int, float]]:
        """
        Get all values of a specific column
        
        Args:
            sheet_name: Name of the sheet
            column_name: Name of the column
            
        Returns:
            List of column values
        """
        data = self.read_excel_file(sheet_name)
        return [row.get(column_name) for row in data if column_name in row and row[column_name] is not None]
